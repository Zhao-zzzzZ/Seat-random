import random
import tkinter as tk
from tkinter import messagebox, simpledialog, filedialog
import json
import os
import sys
import datetime
try:
    import openpyxl
    from openpyxl.styles import Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# 获取资源路径，兼容PyInstaller打包后的情况
def 获取资源路径(相对路径):
    """获取资源文件的绝对路径，兼容开发环境和打包后的环境"""
    if getattr(sys, 'frozen', False):
        # 如果是打包后的环境
        基础路径 = sys._MEIPASS
    else:
        # 如果是开发环境
        基础路径 = os.path.dirname(os.path.abspath(__file__))
    
    return os.path.join(基础路径, 相对路径)

class 座位分配:
    """班级座位随机分配系统主类
    
    功能：
    - 提供图形界面进行座位随机分配
    - 支持特殊座位安排设置
    - 支持导出座位表到Excel
    - 支持管理员密码保护设置功能
    
    主要方法：
    - 随机分配座位(): 执行随机座位分配算法
    - 导出到Excel(): 将当前座位表导出为Excel文件
    - 设置指定排数(): 设置学生必须坐在指定排数
    - 清除设置(): 清除所有特殊安排设置
    """
    def __init__(self, root):
        """初始化座位分配系统
        
        参数:
            root: tkinter根窗口对象
        """
        self.root = root
        self.root.title("随机座位分配")
        self.root.resizable(True, True)
        
        # 设置学生名单
        self.学生名单 = self.加载学生名单()
        
        # 座位布局：6列，左右两列各5人，中间4列各6人
        self.座位行数 = 6  # 最多的列有6行
        self.座位列数 = 6  # 总共6列
        
        # 初始化特殊安排
        self.指定排数安排 = {}  # 格式: {学生: 排数列表}
        
        # 记录当前分配结果
        self.当前分配结果 = {}  # 学生 -> (行, 列)
        self.座位到学生 = {}  # (行, 列) -> 学生

        # 缓存所有有效座位，避免重复计算
        self.有效座位 = [
            (i, j)
            for i in range(self.座位行数)
            for j in range(self.座位列数)
            if not ((j == 0 or j == 5) and i == 5)
        ]
        
        # 用于记录点击状态
        self.第一次点击 = None  # 记录第一次点击的学生和位置
        self.第二次点击 = None  # 记录第二次点击的学生和位置
        
        # 加载特殊安排
        self.加载特殊安排()
        
        # 加载管理员密码
        self.管理员密码 = self.加载管理员密码()
        
        # 创建UI元素
        self.创建界面()
        self.自适应窗口大小()
        
        # 绑定快捷键
        self.root.bind("<Control-Alt-s>", self.显示设置按钮)
    
    def 创建界面(self):
        """创建 Win11 风格界面"""
        self.root.configure(bg="#F3F3F3")

        # 统一样式参数
        self.主背景色 = "#F3F3F3"
        self.卡片背景色 = "#FFFFFF"
        self.主强调色 = "#0078D4"
        self.次要文字色 = "#5F5F5F"
        self.座位默认背景 = "#FFFFFF"
        self.座位选中背景 = "#DCEBFA"
        self.座位卡片宽度 = 92
        self.座位卡片高度 = 54
        self.标题字体 = ("Segoe UI", 18, "bold")
        self.正文粗体 = ("Segoe UI", 10, "bold")
        self.正文常规 = ("Segoe UI", 10)
        self.座位字体 = ("Segoe UI", 9)

        # 顶部标题区域
        顶部框架 = tk.Frame(self.root, bg=self.主背景色)
        顶部框架.pack(fill=tk.X, padx=16, pady=(14, 8))

        标题标签 = tk.Label(
            顶部框架,
            text="班级座位随机分配系统",
            font=self.标题字体,
            bg=self.主背景色,
            fg="#202020"
        )
        标题标签.pack(anchor="w")

        副标题标签 = tk.Label(
            顶部框架,
            text="Windows 11 风格 · 清晰布局 · 高效操作",
            font=("Segoe UI", 9),
            bg=self.主背景色,
            fg=self.次要文字色
        )
        副标题标签.pack(anchor="w", pady=(2, 0))

        # 控制区卡片
        self.控制面板 = tk.Frame(self.root, bg=self.卡片背景色, bd=0, highlightthickness=1, highlightbackground="#E6E6E6")
        self.控制面板.pack(fill=tk.X, padx=16, pady=(0, 10), ipady=8)

        self.随机分配按钮 = self.创建按钮(self.控制面板, "随机分配座位", self.随机分配座位, 主要=True)
        self.随机分配按钮.grid(row=0, column=0, padx=(12, 8), pady=6)

        if EXCEL_AVAILABLE:
            self.导出按钮 = self.创建按钮(self.控制面板, "导出 Excel", self.导出到Excel)
            self.导出按钮.grid(row=0, column=1, padx=8, pady=6)

        self.设置排数按钮 = self.创建按钮(self.控制面板, "设置指定排数", self.设置指定排数)
        self.清除设置按钮 = self.创建按钮(self.控制面板, "清除所有设置", self.清除设置, 危险=True)

        # 座位区域卡片
        self.座位框架 = tk.Frame(self.root, bg=self.卡片背景色, bd=0, highlightthickness=1, highlightbackground="#E6E6E6")
        self.座位框架.pack(fill=tk.BOTH, expand=True, padx=16, pady=(0, 10))
        self.座位内容框架 = tk.Frame(self.座位框架, bg=self.卡片背景色)
        self.座位内容框架.pack(anchor="center", pady=(10, 10))

        讲台标签 = tk.Label(
            self.座位内容框架,
            text="讲台",
            font=self.正文粗体,
            bg="#F5F9FF",
            fg="#1F3A5B",
            relief="flat",
            width=12,
            height=1
        )
        讲台标签.grid(row=0, column=2, columnspan=2, pady=(10, 8))

        self.座位标签 = []
        for i in range(self.座位行数):
            行标签 = []
            for j in range(self.座位列数):
                if (j == 0 or j == 5) and i == 5:
                    标签 = tk.Label(self.座位内容框架, text="", bg=self.卡片背景色)
                else:
                    标签 = self.创建座位标签(self.座位内容框架, i, j)
                    标签.master.grid(row=i + 1, column=j, padx=4, pady=3)
                    标签.bind("<Button-1>", lambda e, row=i, col=j: self.处理座位点击(row, col))
                行标签.append(标签)
            self.座位标签.append(行标签)

        窗户标签 = tk.Label(self.座位内容框架, text="窗户", font=self.正文常规, bg=self.卡片背景色, fg=self.次要文字色)
        窗户标签.grid(row=self.座位行数 + 1, column=0, pady=(6, 10))

        门标签 = tk.Label(self.座位内容框架, text="门", font=self.正文常规, bg=self.卡片背景色, fg=self.次要文字色)
        门标签.grid(row=self.座位行数 + 1, column=self.座位列数 - 1, pady=(6, 10))

        状态栏 = tk.Frame(self.root, bg="#FFFFFF", bd=0, highlightthickness=1, highlightbackground="#E6E6E6")
        状态栏.pack(side=tk.BOTTOM, fill=tk.X, padx=16, pady=(0, 14))

        self.状态标签 = tk.Label(状态栏, text="准备就绪", font=("Segoe UI", 9), bg="#FFFFFF", fg=self.次要文字色)
        self.状态标签.pack(side=tk.RIGHT, padx=12, pady=6)

    def 自适应窗口大小(self):
        """根据当前界面内容设置合适的初始窗口大小，并保留可调整能力"""
        self.root.update_idletasks()
        目标宽度 = max(self.root.winfo_reqwidth() + 24, 760)
        目标高度 = max(self.root.winfo_reqheight() + 24, 640)
        self.root.geometry(f"{目标宽度}x{目标高度}")
        self.root.minsize(目标宽度, 目标高度)

    def 创建座位标签(self, parent, row, col):
        """创建固定尺寸的座位卡片，避免因字体或文字长度导致外框尺寸变化"""
        卡片 = tk.Frame(
            parent,
            width=self.座位卡片宽度,
            height=self.座位卡片高度,
            bg=self.座位默认背景,
            highlightthickness=1,
            highlightbackground="#E5E5E5",
            highlightcolor="#E5E5E5",
            bd=0
        )
        卡片.grid_propagate(False)
        卡片.pack_propagate(False)

        标签 = tk.Label(
            卡片,
            text="空座位",
            relief="flat",
            borderwidth=0,
            font=self.座位字体,
            bg=self.座位默认背景,
            fg="#3A3A3A",
            anchor="center",
            justify="center"
        )
        标签.pack(fill=tk.BOTH, expand=True)

        卡片.bind("<Button-1>", lambda e, r=row, c=col: self.处理座位点击(r, c))
        return 标签

    def 更新座位显示(self, row, col, text, *, 已分配=False, 选中=False):
        """统一渲染座位文字与背景，保持外框尺寸和内部样式一致"""
        背景色 = self.座位选中背景 if 选中 else self.座位默认背景
        字体 = self.正文粗体 if 已分配 else self.座位字体
        文字颜色 = "#202020" if 已分配 else "#3A3A3A"
        标签 = self.座位标签[row][col]
        标签.config(text=text, font=字体, fg=文字颜色, bg=背景色)
        标签.master.config(bg=背景色)

    def 创建按钮(self, parent, text, command, 主要=False, 危险=False):
        """创建统一风格按钮"""
        if 主要:
            bg, fg = self.主强调色, "white"
            active_bg, active_fg = "#106EBE", "white"
        elif 危险:
            bg, fg = "#C42B1C", "white"
            active_bg, active_fg = "#A4262C", "white"
        else:
            bg, fg = "#FFFFFF", "#1F1F1F"
            active_bg, active_fg = "#F3F3F3", "#1F1F1F"

        return tk.Button(
            parent,
            text=text,
            command=command,
            font=self.正文常规,
            relief="flat",
            bd=0,
            padx=14,
            pady=6,
            bg=bg,
            fg=fg,
            activebackground=active_bg,
            activeforeground=active_fg,
            cursor="hand2"
        )
    
    def 随机分配座位(self):
        """执行随机座位分配算法
        
        功能:
        - 根据座位布局和特殊安排随机分配座位
        - 优先满足有特殊安排的学生
        - 尝试最多100次分配，直到满足所有条件
        - 更新UI显示分配结果
        
        返回:
            无返回值，但会更新UI显示和当前分配结果
        """
        座位列表 = self.有效座位
        
        # 检查特殊安排是否可行 - 确保特殊安排不会超过可用座位数
        验证结果, 错误信息 = self.验证特殊安排()
        if not 验证结果:
            messagebox.showerror("错误", f"特殊安排无法满足: {错误信息}\n请修改后重试")
            return
        
        # 尝试多次分配 - 由于随机性，可能需要多次尝试才能满足所有特殊安排
        最大尝试次数 = 100
        for _ in range(最大尝试次数):
            # 复制学生名单和座位列表 - 每次尝试都从原始状态开始
            剩余学生 = self.学生名单.copy()
            剩余座位 = 座位列表.copy()
            分配结果 = {}  # 学生 -> (行, 列)
            
            # 先处理有指定排数的学生 - 确保特殊安排优先满足
            for 学生, 排数列表 in self.指定排数安排.items():
                if 学生 not in 剩余学生:
                    continue  # 学生可能已被分配
                
                # 找出指定排数的所有可用座位
                可用座位 = [座位 for 座位 in 剩余座位 if 座位[0] in 排数列表]
                if not 可用座位:
                    continue  # 没有可用座位，跳过此学生
                
                # 随机选择一个座位并分配
                座位 = random.choice(可用座位)
                分配结果[学生] = 座位
                剩余座位.remove(座位)
                剩余学生.remove(学生)
            
            # 随机分配剩余学生 - 无特殊安排的学生随机分配
            random.shuffle(剩余学生)
            random.shuffle(剩余座位)
            for 学生, 座位 in zip(剩余学生, 剩余座位):
                分配结果[学生] = 座位
            
            # 检查是否成功分配所有学生
            if len(分配结果) == len(self.学生名单):
                # 清理所有有效座位并重置背景色
                for i, j in self.有效座位:
                    self.更新座位显示(i, j, "空座位")
                
                # 更新UI显示 - 在座位标签上显示学生姓名
                for 学生, (行, 列) in 分配结果.items():
                    self.更新座位显示(行, 列, 学生, 已分配=True)
                
                # 保存当前分配结果 - 用于后续导出操作
                self.当前分配结果 = 分配结果.copy()
                self.座位到学生 = {座位: 学生 for 学生, 座位 in 分配结果.items()}
                
                return  # 分配成功，退出方法
        
        # 所有尝试都失败后显示错误
        messagebox.showerror("错误", "无法满足所有特殊安排，请减少限制条件后重试")
    
    def 显示设置按钮(self, event=None):
        """按下Ctrl+Alt+S时显示设置按钮"""
        密码 = simpledialog.askstring("验证", "请输入管理员密码:", show="*")
        if 密码 == self.管理员密码:
            self.设置排数按钮.grid(row=0, column=2, padx=5)
            self.清除设置按钮.grid(row=0, column=3, padx=5)
            messagebox.showinfo("成功", "设置按钮已显示")
        else:
            messagebox.showerror("错误", "密码错误")
    
    def 加载学生名单(self):
        """从学生名单文件加载学生列表
        
        功能:
        - 仅从学生名单.json文件加载学生列表
        - 文件不存在、为空或格式错误时直接报错
        
        返回:
            list: 学生名单列表
        """
        学生名单文件 = "学生名单.json"

        # 尝试获取打包后的路径
        try:
            学生名单路径 = 获取资源路径(学生名单文件)
        except:
            学生名单路径 = 学生名单文件

        if not os.path.exists(学生名单路径):
            messagebox.showerror("错误", f"未找到学生名单文件：\n{学生名单路径}")
            raise FileNotFoundError(f"未找到学生名单文件: {学生名单路径}")

        try:
            with open(学生名单路径, "r", encoding="utf-8") as f:
                数据 = json.load(f)
        except json.JSONDecodeError as e:
            messagebox.showerror("错误", f"学生名单文件格式错误: {str(e)}")
            raise ValueError("学生名单文件格式错误") from e
        except PermissionError as e:
            messagebox.showerror("错误", "没有权限读取学生名单文件")
            raise PermissionError("没有权限读取学生名单文件") from e
        except Exception as e:
            messagebox.showerror("错误", f"加载学生名单失败: {str(e)}")
            raise RuntimeError("加载学生名单失败") from e

        if not isinstance(数据, list) or not 数据:
            messagebox.showerror("错误", "学生名单文件必须是非空数组")
            raise ValueError("学生名单文件必须是非空数组")

        if not all(isinstance(学生, str) and 学生.strip() for 学生 in 数据):
            messagebox.showerror("错误", "学生名单文件中的每一项都必须是非空姓名字符串")
            raise ValueError("学生名单文件中的每一项都必须是非空姓名字符串")

        return [学生.strip() for 学生 in 数据]
            
    def 加载特殊安排(self):
        """从JSON文件加载特殊安排
        
        功能:
        - 尝试从特殊安排.json文件加载特殊座位安排
        - 如果文件不存在或格式错误，则使用空字典
        - 兼容PyInstaller打包后的环境
        
        返回:
            无返回值，但会更新self.指定排数安排
        """
        特殊安排文件 = "特殊安排.json"
        
        # 尝试获取打包后的路径
        try:
            特殊安排路径 = 获取资源路径(特殊安排文件)
        except:
            特殊安排路径 = 特殊安排文件
            
        if not os.path.exists(特殊安排路径):
            # 如果文件不存在，创建一个空的特殊安排文件
            self.保存特殊安排()
            return
        
        try:
            with open(特殊安排路径, "r", encoding="utf-8") as f:
                try:
                    数据 = json.load(f)
                    
                    # 清除现有安排
                    self.指定排数安排 = {}
                    
                    # 加载指定排数安排
                    for 学生, 排数列表 in 数据.get("指定排数安排", {}).items():
                        self.指定排数安排[学生] = 排数列表
                    
                except json.JSONDecodeError as e:
                    messagebox.showerror("错误", f"特殊安排文件格式错误: {str(e)}")
                    return
        except FileNotFoundError:
            # 文件不存在，创建一个空的特殊安排文件
            self.保存特殊安排()
            return
        except PermissionError:
            messagebox.showerror("错误", "没有权限读取特殊安排文件")
            return
        except Exception as e:
            messagebox.showerror("错误", f"加载特殊安排失败: {str(e)}")
            return
    
    def 保存特殊安排(self):
        """保存特殊安排到JSON文件"""
        数据 = {
            "指定排数安排": self.指定排数安排
        }
        
        特殊安排文件 = "特殊安排.json"
        
        try:
            # 尝试获取打包后的路径
            try:
                特殊安排路径 = 获取资源路径(特殊安排文件)
            except:
                特殊安排路径 = 特殊安排文件
            
            with open(特殊安排路径, "w", encoding="utf-8") as f:
                json.dump(数据, f, ensure_ascii=False, indent=4)
        except Exception as e:
            messagebox.showerror("错误", f"保存特殊安排失败: {str(e)}")
    
    def 设置指定排数(self):
        """设置学生坐在指定排数"""
        输入 = simpledialog.askstring("设置指定排数", "请输入学生姓名和排数(用空格分隔，排数从0开始):")
        if not 输入:
            return
        
        输入列表 = 输入.split()
        if len(输入列表) < 2:
            messagebox.showerror("错误", "请输入学生姓名和至少一个排数")
            return
        
        学生 = 输入列表[0]
        排数列表 = []
        
        # 验证学生是否在名单中
        if 学生 not in self.学生名单:
            messagebox.showerror("错误", "学生不在名单中")
            return
        
        # 解析排数
        try:
            for 排数 in 输入列表[1:]:
                排数 = int(排数)
                if 排数 < 0 or 排数 >= self.座位行数:
                    messagebox.showerror("错误", f"排数必须在0到{self.座位行数-1}之间")
                    return
                排数列表.append(排数)
        except ValueError:
            messagebox.showerror("错误", "排数必须是数字")
            return
        
        # 添加指定排数安排
        self.指定排数安排[学生] = 排数列表
        
        # 保存设置
        self.保存特殊安排()
        messagebox.showinfo("成功", f"已设置{学生}坐在第{','.join(map(str, 排数列表))}排")
    
    def 清除设置(self):
        """清除所有特殊安排"""
        # 清除指定排数安排
        self.指定排数安排 = {}
        
        # 清除座位显示
        for i in range(self.座位行数):
            for j in range(self.座位列数):
                if not ((j == 0 or j == 5) and i == 5):
                    self.更新座位显示(i, j, "空座位")

        # 清空当前分配缓存
        self.当前分配结果 = {}
        self.座位到学生 = {}
        
        # 保存设置
        self.保存特殊安排()
        messagebox.showinfo("成功", "已清除所有设置")
    
    def 验证特殊安排(self):
        """验证特殊座位安排是否可行
        
        功能:
        - 检查指定排数的学生数量是否超过总座位数
        - 确保特殊安排不会导致座位不足
        
        注意:
        - 当前实现仅检查学生数量是否超过总座位数
        - 不检查排数有效性(由设置指定排数方法处理)
        
        返回:
            tuple: (验证结果, 错误信息)
            - 验证结果: True表示验证通过，False表示验证失败
            - 错误信息: 验证失败时的详细错误描述
        """
        # 计算指定排数的学生数量
        指定排数学生数 = len(self.指定排数安排)
        
        # 计算总座位数 - 减去两个角落没有的座位
        总座位数 = len(self.有效座位)
        
        # 检查学生数量是否超过可用座位数
        if 指定排数学生数 > 总座位数:
            return False, f"指定排数的学生数量({指定排数学生数})超过可用座位数({总座位数})"
        
        return True, ""  # 验证通过

    def 加载管理员密码(self):
        """从配置文件加载管理员密码"""
        try:
            with open("配置.json", "r", encoding="utf-8") as f:
                配置 = json.load(f)
                return 配置.get("管理员密码", "admin")  # 默认密码为admin
        except:
            return "admin"  # 如果配置文件不存在，使用默认密码

    def 导出到Excel(self):
        """导出当前座位表到Excel文件
        
        功能:
        - 创建包含两个工作表的Excel文件:
          1. 座位表(学生视角): 从学生角度看的座位布局
          2. 座位表(讲台视角): 从讲台角度看的座位布局(行列翻转)
        - 添加标题、讲台标识和方向标识
        - 设置单元格格式(居中、边框等)
        - 自动生成带时间戳的文件名
        
        返回:
            无返回值，但会显示导出成功或失败的提示信息
        """
        if not self.当前分配结果:
            messagebox.showerror("错误", "请先进行座位分配")
            return
            
        # 创建新的工作簿
        wb = openpyxl.Workbook()
        
        # 创建第一个工作表（正常视图）- 学生视角
        ws1 = wb.active
        ws1.title = "座位表（学生视角）"
        
        # 创建第二个工作表（翻转视图）- 讲台视角
        ws2 = wb.create_sheet("座位表（讲台视角）")
        
        # 设置列宽 - 统一所有列的宽度为15个字符
        for col in range(1, self.座位列数 + 1):
            ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
            ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # 添加标题（仅学生视角）- 合并第一行的所有列
        ws1.merge_cells(f'A1:{openpyxl.utils.get_column_letter(self.座位列数)}1')
        标题单元格 = ws1.cell(1, 1, "班级座位表")
        标题单元格.alignment = Alignment(horizontal='center', vertical='center')
        标题单元格.font = openpyxl.styles.Font(size=14, bold=True)
        
        # 添加讲台（学生视角 - 顶部）- 合并C2和D2单元格
        ws1.merge_cells(f'C2:D2')
        讲台单元格 = ws1.cell(2, 3, "讲台")
        讲台单元格.alignment = Alignment(horizontal='center', vertical='center')
        讲台单元格.font = openpyxl.styles.Font(bold=True)
        
        # 添加座位（正常视图）- 遍历所有分配结果
        for 学生, (行, 列) in self.当前分配结果.items():
            # 调整行号（Excel从1开始，且第1行是标题，第2行是讲台）
            excel行 = 行 + 3
            excel列 = 列 + 1
            
            # 设置单元格值（正常视图）- 学生姓名居中显示
            单元格 = ws1.cell(excel行, excel列, 学生)
            单元格.alignment = Alignment(horizontal='center', vertical='center')
            单元格.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 设置单元格值（翻转视图）- 行列位置翻转
            # 计算翻转后的位置：行和列都翻转
            翻转行 = self.座位行数 - 行 - 1
            翻转列 = self.座位列数 - 列 - 1
            翻转excel行 = 翻转行 + 1  # 讲台视角不需要标题和讲台行，所以从第1行开始
            翻转excel列 = 翻转列 + 1
            
            翻转单元格 = ws2.cell(翻转excel行, 翻转excel列, 学生)
            翻转单元格.alignment = Alignment(horizontal='center', vertical='center')
            翻转单元格.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # 添加讲台（讲台视角 - 底部，紧贴座位区域）
        ws2.merge_cells(f'C{self.座位行数 + 1}:D{self.座位行数 + 1}')
        翻转讲台单元格 = ws2.cell(self.座位行数 + 1, 3, "讲台")
        翻转讲台单元格.alignment = Alignment(horizontal='center', vertical='center')
        翻转讲台单元格.font = openpyxl.styles.Font(bold=True)
        
        # 添加方向标识（仅学生视角）- 窗户和门标识
        窗户单元格 = ws1.cell(self.座位行数 + 3, 1, "窗户")
        窗户单元格.font = openpyxl.styles.Font(color="0000FF")
        
        门单元格 = ws1.cell(self.座位行数 + 3, self.座位列数, "门")
        门单元格.font = openpyxl.styles.Font(color="0000FF")
        
        # 保存文件 - 使用当前时间生成文件名
        文件名 = f"座位表_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        文件路径 = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=文件名,
            filetypes=[("Excel文件", "*.xlsx")]
        )
        
        if 文件路径:
            try:
                wb.save(文件路径)
                messagebox.showinfo("成功", f"座位表已导出到：\n{文件路径}\n\n包含两个工作表：\n1. 座位表（学生视角）\n2. 座位表（讲台视角）")
            except Exception as e:
                messagebox.showerror("错误", f"导出失败：{str(e)}")

    def 处理座位点击(self, row, col):
        """处理座位点击事件
        
        功能:
        - 记录第一次和第二次点击的座位
        - 当点击两个座位后，执行座位互换
        - 更新界面显示和状态栏提示
        
        参数:
            row: 点击的座位行号
            col: 点击的座位列号
        """
        # 如果还没有进行座位分配，直接返回
        if not self.当前分配结果:
            self.状态标签.config(text="请先进行座位分配")
            return
            
        # 获取点击的座位上的学生
        当前学生 = self.座位到学生.get((row, col))
                
        if not 当前学生:
            self.状态标签.config(text="请点击有学生的座位")
            return
            
        # 如果是第一次点击
        if self.第一次点击 is None:
            self.第一次点击 = (当前学生, row, col)
            self.状态标签.config(text=f"已选择{当前学生}，请选择要交换的学生")
            # 高亮显示选中的座位
            self.更新座位显示(row, col, 当前学生, 已分配=True, 选中=True)
            return
            
        # 如果是第二次点击
        if self.第二次点击 is None:
            # 如果点击的是同一个座位
            if self.第一次点击[0] == 当前学生:
                self.状态标签.config(text="请选择不同的学生进行交换")
                return
                
            self.第二次点击 = (当前学生, row, col)
            self.状态标签.config(text=f"正在交换{self.第一次点击[0]}和{当前学生}的座位")
            
            # 执行座位互换
            self.互换座位()
            
            # 重置点击状态
            self.第一次点击 = None
            self.第二次点击 = None
            
    def 互换座位(self):
        """执行座位互换操作
        
        功能:
        - 交换两个学生的座位位置
        - 更新界面显示
        - 更新当前分配结果
        """
        if not (self.第一次点击 and self.第二次点击):
            return
            
        学生1, 行1, 列1 = self.第一次点击
        学生2, 行2, 列2 = self.第二次点击
        
        # 更新当前分配结果
        self.当前分配结果[学生1] = (行2, 列2)
        self.当前分配结果[学生2] = (行1, 列1)
        self.座位到学生[(行1, 列1)] = 学生2
        self.座位到学生[(行2, 列2)] = 学生1
        
        # 更新界面显示
        self.更新座位显示(行1, 列1, 学生2, 已分配=True)
        self.更新座位显示(行2, 列2, 学生1, 已分配=True)
        
        # 重置所有座位的背景色
        for i, j in self.有效座位:
            if (i, j) in self.座位到学生:
                self.更新座位显示(i, j, self.座位到学生[(i, j)], 已分配=True)
            else:
                self.更新座位显示(i, j, "空座位")
        
        # 更新状态栏
        self.状态标签.config(text=f"已成功交换{学生1}和{学生2}的座位")

if __name__ == "__main__":
    root = tk.Tk()
    try:
        app = 座位分配(root)
    except Exception:
        root.destroy()
        raise SystemExit(1)
    root.mainloop()
